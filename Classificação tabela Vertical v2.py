# Requisitos: pandas, openpyxl
# pip install pandas openpyxl

import pandas as pd
import os, re
from datetime import datetime
from collections import OrderedDict
from openpyxl import load_workbook
from tkinter import (
    Tk, Label, Button, filedialog, StringVar, OptionMenu, messagebox,
    Checkbutton, BooleanVar, Frame, Scrollbar, Canvas,
    Listbox, SINGLE, END, Entry
)

# ---------- Config ----------#
OUT_DIR = "LOG_files"
os.makedirs(OUT_DIR, exist_ok=True)
DEFAULT_OUTPUT = "Classificação Gerada.xlsx"

# ---------- Helpers ----------#
def detect_decimal_info(s):
    if s is None: return (None, 0)
    s0 = str(s).strip()
    if s0 == "": return (None, 0)
    s0 = s0.replace(" ", "")
    try:
        if "." in s0 and "," in s0:
            if s0.rfind(".") < s0.rfind(","):
                snum = s0.replace(".", "").replace(",", ".")
            else:
                snum = s0.replace(",", "")
        elif "," in s0:
            snum = s0.replace(",", ".")
        else:
            snum = s0
        val = float(snum)
        dec = len(snum.split(".")[1].rstrip("0")) if "." in snum else 0
        return (val, dec)
    except:
        return (None, 0)

def number_format_for_decimals(d):
    if d <= 0:
        return "#,##0"
    return "#,##0." + ("0"*d)

def preserve(v):
    return "" if pd.isna(v) else str(v).strip()

def to_num_for_sort(s):
    try:
        if s is None or str(s).strip() == "":
            return float("inf")
        return float(str(s).replace(",","."))
    except:
        return float("inf")

def find_col(df_cols, aliases):
    for a in aliases:
        for c in df_cols:
            if isinstance(c, str) and c.strip().lower() == a.strip().lower():
                return c
    for c in df_cols:
        if not isinstance(c, str): continue
        cl = re.sub(r'[^a-z0-9]', '', c.lower())
        for a in aliases:
            an = re.sub(r'[^a-z0-9]', '', a.lower())
            if an in cl:
                return c
    return None

# ---------- Dark helpers ----------#
def apply_dark_background_to_frames(root_widget, bg="#1e1e1e"):
    try:
        root_widget.configure(bg=bg)
    except:
        pass
    stack = [root_widget]
    while stack:
        w = stack.pop()
        for child in w.winfo_children():
            stack.append(child)
            try:
                if child.winfo_class() == "Frame":
                    child.configure(bg=bg)
                if child.winfo_class() == "TFrame":
                    try:
                        child.configure(background=bg)
                    except:
                        pass
            except:
                pass

def darken_input_widgets(root,
                         entry_bg="#f6f6f6", entry_fg="#000000",
                         listbox_bg="#2b2b2b", listbox_fg="#eaeaea",
                         canvas_bg="#f0f0f0"):
    stack = [root]
    while stack:
        w = stack.pop()
        for child in w.winfo_children():
            stack.append(child)
            try:
                cls = child.winfo_class()
            except:
                continue

            if cls in ("Entry", "TEntry"):
                try:
                    child.configure(bg=entry_bg, fg=entry_fg, insertbackground=entry_fg)
                except:
                    pass
            elif cls == "Listbox":
                try:
                    child.configure(bg=listbox_bg, fg=listbox_fg,
                                    selectbackground="#555555", selectforeground=listbox_fg)
                except:
                    pass
            elif cls == "Canvas":
                try:
                    child.configure(bg=canvas_bg)
                except:
                    pass
            elif cls == "Text":
                try:
                    child.configure(bg=entry_bg, fg=entry_fg, insertbackground=entry_fg)
                except:
                    pass
            elif cls in ("Menubutton",):
                try:
                    child.configure(bg=entry_bg, fg=entry_fg)
                except:
                    pass

# ---------- Core processing ----------#
def process_file_custom_token(input_path, output_path,
                 col_weight_from, col_weight_to, col_uf, col_city,
                 token_columns, col_zip_from, col_zip_to, col_time):
    print("Início:", datetime.now())
    df = pd.read_excel(input_path, dtype=str)
    df.fillna("", inplace=True)
    print("Arquivo lido:", input_path, "linhas:", len(df))

    if col_uf and col_uf in df.columns:
        df[col_uf] = df[col_uf].astype(str).fillna("").str.strip().str.upper()
    else:
        df["UF_TMP"] = ""
        col_uf = "UF_TMP"

    if col_city and col_city in df.columns:
        df[col_city] = df[col_city].astype(str).fillna("").str.strip()

    df["_BUCKET"] = df.apply(lambda r: f"{preserve(r.get(col_weight_from,''))}|{preserve(r.get(col_weight_to,''))}", axis=1)

    def build_token_row(r):
        parts = []
        for c in token_columns:
            parts.append(preserve(r.get(c, "")))
        return "|".join(parts) if parts else ""
    df["_TOKEN"] = df.apply(build_token_row, axis=1)

    if col_zip_from and col_zip_to and (col_zip_from in df.columns) and (col_zip_to in df.columns):
        df["CONCAT_CEP"] = (df[col_zip_from].astype(str).str.replace(r'\.0+$','',regex=True).str.strip() +
                            df[col_zip_to].astype(str).str.replace(r'\.0+$','',regex=True).str.strip())
    else:
        df["CONCAT_CEP"] = df.apply(lambda r: preserve(r.get(col_zip_from,"")) + preserve(r.get(col_zip_to,"")), axis=1)

    if col_weight_from not in df.columns:
        unique_buckets = df[["_BUCKET"]].drop_duplicates().copy()
        unique_buckets[col_weight_from if col_weight_from else "_wf_dummy"] = ""
    else:
        unique_buckets = df[["_BUCKET", col_weight_from]].drop_duplicates().copy()
    unique_buckets["_wfnum"] = unique_buckets[col_weight_from].apply(lambda x: to_num_for_sort(x) if col_weight_from in unique_buckets.columns else float("inf"))
    unique_buckets_sorted = unique_buckets.sort_values("_wfnum")["_BUCKET"].tolist()

    index_cols = [col_uf, col_city] if (col_city and col_city in df.columns) else [col_uf, "CONCAT_CEP"]
    pivot = df.pivot_table(index=index_cols, columns="_BUCKET", values="_TOKEN", aggfunc=lambda vals: ";".join(map(str, vals)))
    for b in unique_buckets_sorted:
        if b not in pivot.columns:
            pivot[b] = ""
    pivot = pivot[unique_buckets_sorted]
    pivot["_FINGERPRINT"] = pivot.apply(lambda r: "|".join([str(r[b]) for b in unique_buckets_sorted]), axis=1)

    uf_seq = {}
    key_to_class = {}
    for (uf, ident), row in pivot.iterrows():
        fp = row["_FINGERPRINT"]
        if uf not in uf_seq:
            uf_seq[uf] = OrderedDict()
        if fp not in uf_seq[uf]:
            uf_seq[uf][fp] = len(uf_seq[uf]) + 1
        seq = uf_seq[uf][fp]
        key_to_class[(uf, ident)] = f"{uf}{seq}"

    def group_key(r):
        if col_city and col_city in df.columns:
            return (r.get(col_uf,""), r.get(col_city,""))
        return (r.get(col_uf,""), r.get("CONCAT_CEP",""))
    df["_GROUP_KEY"] = df.apply(group_key, axis=1)
    df["CLASS"] = df["_GROUP_KEY"].map(key_to_class)

    for c in token_columns:
        if c not in df.columns:
            df[c] = ""
    price_df = df[[col_uf, "_BUCKET"] + token_columns + ["CLASS"]].drop_duplicates().reset_index(drop=True)
    price_df[["WeightFrom","WeightTo"]] = price_df["_BUCKET"].str.split("|", expand=True)

    decimals_map = {}
    for c in token_columns:
        decs = []
        for v in price_df[c].astype(str).head(500):
            _, d = detect_decimal_info(v)
            decs.append(d)
        decimals_map[c] = max(decs) if decs else 0

    price_out = price_df[["CLASS","WeightFrom","WeightTo"] + token_columns].copy()

    coverage_cols = [col_uf]
    if col_city and col_city in df.columns: coverage_cols.append(col_city)
    if col_time and col_time in df.columns: coverage_cols.append(col_time)
    if col_zip_from and col_zip_from in df.columns: coverage_cols.append(col_zip_from)
    if col_zip_to and col_zip_to in df.columns: coverage_cols.append(col_zip_to)
    coverage_cols.extend(["CONCAT_CEP","CLASS"])
    coverage = df[coverage_cols].drop_duplicates().reset_index(drop=True)

    price_df.to_csv(os.path.join(OUT_DIR, "Preco_raw_preview.csv"), index=False, sep=";", encoding="utf-8-sig")
    coverage.to_csv(os.path.join(OUT_DIR, "Prazo_raw_preview.csv"), index=False, sep=";", encoding="utf-8-sig")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        price_out.to_excel(writer, sheet_name="Preço", index=False)
        coverage.to_excel(writer, sheet_name="Prazo", index=False)

    wb = load_workbook(output_path)
    ws = wb["Preço"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: i+1 for i, name in enumerate(header)}

    for token_col in token_columns:
        if token_col in col_idx:
            dec = decimals_map.get(token_col, 0)
            fmt = number_format_for_decimals(dec)
            for row in ws.iter_rows(min_row=2, min_col=col_idx[token_col], max_col=col_idx[token_col], max_row=ws.max_row):
                cell = row[0]
                if cell.value is None or cell.value == "":
                    continue
                if isinstance(cell.value, str):
                    parsed, _ = detect_decimal_info(cell.value)
                    if parsed is not None:
                        cell.value = parsed
                cell.number_format = fmt

    wb.save(output_path)
    print("Concluído:", datetime.now())
    print("Arquivo gerado:", output_path)
    print("Previews em:", OUT_DIR)
    messagebox.showinfo("Concluído", f"Arquivo gerado:\n{output_path}\nPreviews: {OUT_DIR}")

# ---------- GUI ----------#
def open_gui():
    root = Tk()
    root.title("Gerar Classificação")
    root.geometry("1100x820")

    selected_file_var = StringVar()
    selected_file_var.set("Nenhum arquivo selecionado")
    

    map_keys = ["Peso de","Peso até","UF","Cidade","Preço","Preço extra","CEP de","CEP até","Prazo"]
    map_vars = {k: StringVar() for k in map_keys}
    map_vars_menus = {}

    cols = []
    checkbox_vars = {}

    # top area
    top_frame = Frame(root)
    top_frame.pack(fill="x", padx=12)
    btn_color = "#0c005a"
    Button(top_frame, text="Selecionar arquivo", command=lambda: choose_file(root, selected_file_var, map_vars, map_vars_menus, cols, checkbox_vars), bg=btn_color, fg="white").pack(side="left", pady=6)
    label_arquivo = Label(top_frame, textvariable=selected_file_var,wraplength=900, bg="#1e1e1e", fg="red") 
    label_arquivo.pack(side="left", padx=12) 


    # mapping area
    map_frame = Frame(root)
    map_frame.pack(fill="x", padx=12, pady=(8,6))
    Label(map_frame, text="2) Mapeie as colunas principais (deixe em branco as opcionais):", font=("Arial", 11, "bold"), bg="#1e1e1e", fg="#eaeaea").pack(anchor="w")
    maps_inner = Frame(map_frame); maps_inner.pack(fill="x", pady=(6,4))
    left_col = Frame(maps_inner); right_col = Frame(maps_inner)
    left_col.pack(side="left", fill="x", expand=True, padx=(0,6))
    right_col.pack(side="left", fill="x", expand=True, padx=(6,0))

    left_keys = ["Peso de","Peso até","UF","Cidade","Preço"]
    right_keys = ["Preço extra","CEP de","CEP até","Prazo"]
    for k in left_keys:
        Label(left_col, text=k + ":", bg="#1e1e1e", fg="#eaeaea").pack(anchor="w")
        map_vars[k].set("")
        om = OptionMenu(left_col, map_vars[k], "")
        # style OptionMenu (botão e seu menu)
        try:
            om.configure(bg="#f6f6f6", fg="#000000", activebackground="#e8e8e8")
        except:
            pass
        try:
            menu_widget = om["menu"]
            menu_widget.configure(bg="#f6f6f6", fg="#000000", activebackground="#e8e8e8")
        except:
            pass
        om.pack(fill="x", pady=(0,6))
        map_vars_menus[k] = om
    for k in right_keys:
        Label(right_col, text=k + ":", bg="#1e1e1e", fg="#eaeaea").pack(anchor="w")
        map_vars[k].set("")
        om = OptionMenu(right_col, map_vars[k], "")
        try:
            om.configure(bg="#f6f6f6", fg="#000000", activebackground="#e8e8e8")
        except:
            pass
        try:
            menu_widget = om["menu"]
            menu_widget.configure(bg="#f6f6f6", fg="#000000", activebackground="#e8e8e8")
        except:
            pass
        om.pack(fill="x", pady=(0,6))
        map_vars_menus[k] = om

    # lower panel
    lower_panel = Frame(root)
    lower_panel.pack(fill="both", expand=True, padx=12, pady=(6,10))

    # token list (left)
    token_left = Frame(lower_panel); token_left.pack(side="left", fill="y", padx=(0,8))
    Label(token_left, text="Token order (use ↑ ↓ para reordenar):", font=("Arial", 10, "bold"),  bg="#1e1e1e", fg="#eaeaea").pack(anchor="w")
    token_listbox = Listbox(token_left, selectmode=SINGLE, width=40, height=18); token_listbox.pack(pady=(6,8))
    preview_var = StringVar()
    Label(token_left, text="Preview do TOKEN:", font=("Arial", 10, "bold"), bg="#1e1e1e", fg="#eaeaea").pack(anchor="w")
    preview_label = Label(token_left, textvariable=preview_var, wraplength=320, justify="left", anchor="w"); preview_label.pack(fill="x", pady=(4,0))

    # control buttons (middle)
    ctrl_mid = Frame(lower_panel, width=60); ctrl_mid.pack(side="left", fill="y", padx=(4,8)); ctrl_mid.pack_propagate(False)
    Button(ctrl_mid, text="↑", width=3, command=lambda: move_up(token_listbox, preview_var)).pack(pady=(60,6))
    Button(ctrl_mid, text="↓", width=3, command=lambda: move_down(token_listbox, preview_var)).pack(pady=(0,6))
    Button(ctrl_mid, text="Remover", width=8, command=lambda: remove_selected(token_listbox, preview_var)).pack(pady=(8,6))

    # checkbox canvas (right)
    checklist_frame = Frame(lower_panel); checklist_frame.pack(side="left", fill="both", expand=True)
    Label(checklist_frame, text="Marque as colunas que devem entrar no TOKEN por faixa (clique para adicionar/remover):", font=("Arial", 10, "bold"),  bg="#1e1e1e", fg="#eaeaea").pack(anchor="w", padx=4, pady=(0,6))
    canvas_holder = Frame(checklist_frame, bd=1, relief="sunken"); canvas_holder.pack(fill="both", expand=True, padx=2, pady=(0,6))

    # canvas: area clara (onde as checkbuttons vão), scrollbar escura
    canvas = Canvas(canvas_holder)
    try:
        canvas.configure(bg="#f0f0f0", highlightthickness=0)
    except:
        pass
    vsb = Scrollbar(canvas_holder, orient="vertical", command=canvas.yview)
    try:
        vsb.configure(bg="#2b2b2b", troughcolor="#2b2b2b", activebackground="#555555")
    except:
        pass

    inner_frame = Frame(canvas)
    inner_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
    canvas.create_window((0, 0), window=inner_frame, anchor="nw")
    canvas.configure(yscrollcommand=vsb.set)
    canvas.pack(side="left", fill="both", expand=True); vsb.pack(side="right", fill="y")

    btns_frame = Frame(checklist_frame); btns_frame.pack(fill="x", pady=(0,6))
    Button(btns_frame, text="Selecionar tudo", command=lambda: select_all(checkbox_vars, token_listbox, preview_var)).pack(side="left", padx=6)
    Button(btns_frame, text="Desmarcar tudo", command=lambda: deselect_all(checkbox_vars, token_listbox, preview_var)).pack(side="left", padx=6)

    # bottom section
    bottom_frame = Frame(root); bottom_frame.pack(fill="x", padx=12, pady=(10, 16))
    Label(bottom_frame, text="3) Nome do arquivo de saída (opcional):", font=("Arial", 11, "bold"),  bg="#1e1e1e", fg="#eaeaea").pack(anchor="center", pady=(0,6))
    out_name_var = StringVar(); out_name_var.set(DEFAULT_OUTPUT)
    Entry(bottom_frame, textvariable=out_name_var, width=60, justify="center", font=("Arial", 10)).pack(pady=(0,8))
    Button(bottom_frame, text="Gerar Classificação", command=lambda: run_process(selected_file_var, map_vars, token_listbox, out_name_var), bg=btn_color, fg="white", font=("Arial", 11, "bold"), width=25).pack(pady=(8,0))

    # helper functions
    def update_preview_local():
        items = token_listbox.get(0, END)
        preview_var.set(" | ".join(items))

    def move_up(lb, pv):
        sel = lb.curselection()
        if not sel: return
        i = sel[0]
        if i == 0: return
        text = lb.get(i)
        lb.delete(i); lb.insert(i-1, text); lb.select_set(i-1); update_preview_local()

    def move_down(lb, pv):
        sel = lb.curselection()
        if not sel: return
        i = sel[0]
        if i == lb.size()-1: return
        text = lb.get(i)
        lb.delete(i); lb.insert(i+1, text); lb.select_set(i+1); update_preview_local()

    def remove_selected(lb, pv):
        sel = lb.curselection()
        if not sel: return
        lb.delete(sel[0]); update_preview_local()

    def choose_file(root_obj, selected_file_var, map_vars, map_vars_menus, cols_ref, checkbox_vars_ref):
        nonlocal cols, checkbox_vars, inner_frame, canvas
        path = filedialog.askopenfilename(filetypes=[("Excel files","*.xlsx *.xls")])
        if not path: return
        selected_file_var.set(path)
        try:
            df_head = pd.read_excel(path, nrows=1, dtype=str)
            cols = df_head.columns.tolist()
            selected_file_var.set(path)
            label_arquivo.config(fg="lime")
            
        except Exception as e:
            messagebox.showerror("Erro leitura", f"Não foi possível ler o arquivo:\n{e}")
            return

        # populate dropdowns
        for key, var in map_vars.items():
            menu = map_vars_menus[key]
            menu["menu"].delete(0, "end")
            menu["menu"].add_command(label="", command=lambda v=var: v.set(""))
            for c in cols:
                menu["menu"].add_command(label=c, command=lambda value=c, v=var: v.set(value))
            # autodetect
            if key == "Peso de":
                var.set(find_col(cols, ["WeightStart","Peso de","VL_WEIGHT_FROM","WeightFrom","Weight Start","PesoInicio"]) or "")
            elif key == "Peso até":
                var.set(find_col(cols, ["WeightEnd","Peso ate","VL_WEIGHT_TO","WeightTo","Weight End","PesoFim"]) or "")
            elif key == "UF":
                var.set(find_col(cols, ["UF","DC_UF","Estado"]) or "")
            elif key == "Cidade":
                var.set(find_col(cols, ["CIDADE","City","Cidade","Municipio"]) or "")
            elif key == "Preço":
                var.set(find_col(cols, ["AbsoluteMoneyCos","AbsoluteMoneyCost","Valor do frete","VL_SHIPPING","Price","Valor"]) or "")
            elif key == "Preço extra":
                var.set(find_col(cols, ["PriceByExtraWeight","Price By Extra Weight","Price_By_Extra","ExtraWeightPrice","Ad Valorem","AdValorem"]) or "")
            elif key == "CEP de":
                var.set(find_col(cols, ["ZipCodeStart","CEP DE","NM_DESTINATION_CEP_FROM","Cep de"]) or "")
            elif key == "CEP até":
                var.set(find_col(cols, ["ZipCodeEnd","CEP ATE","NM_DESTINATION_CEP_TO","Cep ate"]) or "")
            elif key == "Prazo":
                var.set(find_col(cols, ["TimeCost","PRAZO","VL_DELIVERY_TIME","Time"]) or "")

        # clear previous checkbox area and listbox
        for widget in inner_frame.winfo_children():
            widget.destroy()
        token_listbox.delete(0, END)
        checkbox_vars.clear()

        # create checkbuttons in inner_frame (right panel)
        cb_bg = "#f0f0f0"
        cb_active = "#00f320"
        for c in cols:
            var = BooleanVar(value=False)
            cb = Checkbutton(inner_frame, text=c, variable=var, command=(lambda col=c: on_chk_toggle(col)()), anchor="w", justify="left")
            # set colors explicitly
            try:
                cb.configure(bg=cb_bg, fg="#000000", activebackground=cb_active, selectcolor=cb_bg)
            except:
                pass
            cb.pack(anchor="w", padx=6, pady=2, fill="x")
            checkbox_vars[c] = var

        # helper to add/remove from listbox (we need a closure)
        def on_chk_toggle(colname):
            def _inner():
                if checkbox_vars[colname].get():
                    if colname not in token_listbox.get(0, END):
                        token_listbox.insert(END, colname)
                else:
                    items = token_listbox.get(0, END)
                    if colname in items:
                        idx = items.index(colname)
                        token_listbox.delete(idx)
                update_preview_local()
            return _inner

        # pre-check common names and add to token_listbox
        for guess in ["AbsoluteMoneyCos","AbsoluteMoneyCost","Valor do frete","Price","PriceByExtraWeight","TRT","TDA","VL_TRT","VL_TDA","VL_SHIPPING","Ad Valorem","AdValorem","VL_TOLL_PER_100KG"]:
            for colname in cols:
                if colname.strip().lower() == guess.strip().lower() and colname in checkbox_vars:
                    checkbox_vars[colname].set(True)
                    token_listbox.insert(END, colname)
        update_preview_local()

        # bind mousewheel to canvas only when pointer over it
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        def bind_mousewheel(e):
            canvas.bind("<MouseWheel>", _on_mousewheel)
        def unbind_mousewheel(e):
            canvas.unbind("<MouseWheel>")
        canvas.bind("<Enter>", bind_mousewheel)
        canvas.bind("<Leave>", unbind_mousewheel)

    def select_all(vars_map, lb, pv):
        lb.delete(0, END)
        for k, v in vars_map.items():
            v.set(True)
            lb.insert(END, k)
        update_preview_local()

    def deselect_all(vars_map, lb, pv):
        for k, v in vars_map.items():
            v.set(False)
        lb.delete(0, END)
        update_preview_local()

    def run_process(selected_file_var, map_vars, token_listbox, out_name_var):
        in_path = selected_file_var.get()
        if in_path is None or not os.path.exists(in_path):
            messagebox.showerror("Erro", "Selecione um arquivo válido antes de continuar.")
            return
        mappings = {k: (v.get() if v.get() != "" else None) for k,v in map_vars.items()}
        selected_token_cols = list(token_listbox.get(0, END))
        if not selected_token_cols:
            if not messagebox.askyesno("Confirmar", "Nenhuma coluna marcada para token. Deseja continuar (o token ficará vazio)?"):
                return
        output_name = out_name_var.get().strip() or DEFAULT_OUTPUT
        output_path = os.path.join(os.path.dirname(in_path), output_name)
        try:
            process_file_custom_token(
                in_path, output_path,
                col_weight_from=mappings["Peso de"],
                col_weight_to=mappings["Peso até"],
                col_uf=mappings["UF"],
                col_city=mappings["Cidade"],
                token_columns=selected_token_cols,
                col_zip_from=mappings["CEP de"],
                col_zip_to=mappings["CEP até"],
                col_time=mappings["Prazo"]
            )
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Erro durante processamento", str(e))

    # apply dark background to frames
    apply_dark_background_to_frames(root, bg="#1e1e1e")

    # adjust inputs/listboxes/canvas (entries claras, listbox escuro, canvas claro)
    darken_input_widgets(root,
                         entry_bg="#eaeaea", entry_fg="#000000",
                         listbox_bg="#2b2b2b", listbox_fg="#eaeaea",
                         canvas_bg="#2b2b2b")

    root.mainloop()

if __name__ == "__main__":
    open_gui()
